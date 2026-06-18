from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "2023 Bylaws_3_3_2025_CVL.xlsx"
CONTENT_DIR = ROOT / "src" / "content"

TRACKS = [
    {
        "id": "smart-structures",
        "name": "Smart Structures Track",
        "sourceTrackName": "Smart Structural Engineering (SSE)",
        "sourceSheet": "CCE Major-2 (8 S)",
        "careerDirection": "Structural engineering, smart infrastructure, monitoring, BIM, and resilient built assets.",
    },
    {
        "id": "project-management",
        "name": "Project Management Track",
        "sourceTrackName": "Construction Engineering & Management (CEM)",
        "sourceSheet": "CCE Major-1 (8 S)",
        "careerDirection": "Construction planning, project delivery, contracts, cost control, and infrastructure management.",
    },
]

SEMESTER_BLOCKS = [
    {"year": "Year 1", "titleRow": 12, "startRow": 14, "endRow": 21},
    {"year": "Year 2", "titleRow": 24, "startRow": 26, "endRow": 33},
    {"year": "Year 3", "titleRow": 36, "startRow": 38, "endRow": 45},
    {"year": "Year 4", "titleRow": 48, "startRow": 50, "endRow": 56},
]

LEFT = {
    "termCol": 1,
    "codeCol": 2,
    "nameCol": 3,
    "creditCol": 4,
    "prereqCol": 11,
    "flagCols": {"university": 12, "eas": 13, "basicScience": 14, "program": 15, "major": 16, "elective": 17},
}
RIGHT = {
    "termCol": 18,
    "codeCol": 19,
    "nameCol": 20,
    "creditCol": 21,
    "prereqCol": 28,
    "flagCols": {"university": 29, "eas": 30, "basicScience": 31, "program": 32, "major": 33, "elective": 34},
}

ZERO_CREDIT_CODES = {"EGL111", "EGL112", "CVL391", "CVL392"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def as_int(value: Any, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def assign_credits(code: str, name: str, workbook_value: Any) -> tuple[int, str | None]:
    workbook_credit = as_int(workbook_value)
    if code in ZERO_CREDIT_CODES:
        assigned = 0
    elif "Writing Skills" in name or "Presentation Skills" in name:
        assigned = 3
    elif workbook_credit is not None:
        assigned = workbook_credit
    else:
        assigned = 3

    note = None
    if workbook_credit is not None and workbook_credit != assigned:
        note = f"{code} {name}: workbook value {workbook_credit} overridden to {assigned} by explicit rule."
    if workbook_credit is None and assigned == 3:
        note = f"{code} {name}: no workbook value; defaulted to 3 credit hours."
    return assigned, note


def read_flags(ws, row: int, block: dict[str, Any]) -> dict[str, int]:
    return {name: as_int(ws.cell(row, col).value, 0) or 0 for name, col in block["flagCols"].items()}


def course_type(code: str, name: str, flags: dict[str, int]) -> str:
    lower = name.lower()
    if "practical training" in lower:
        return "practical"
    if "elective" in lower or code in {"MTH E", "CVL E1", "CVL E2", "CVL E3", "CVL E4", "HUM E1", "INT E1", "INT E2"}:
        return "elective"
    if flags.get("university"):
        return "university"
    if flags.get("major"):
        return "track"
    return "core"


def course_profile(name: str, kind: str, track_name: str) -> dict[str, Any]:
    lower = name.lower()
    track_careers = {
        "Smart Structures Track": ["Structural design engineer", "BIM engineer", "Smart infrastructure engineer"],
        "Project Management Track": ["Planning engineer", "Project controls engineer", "Construction project engineer"],
    }[track_name]

    if kind == "practical":
        return {
            "description": "A zero-credit field training requirement that links classroom study with supervised civil engineering practice.",
            "skills": ["Site observation", "Field documentation", "Professional communication"],
            "jobMarket": "Employers value graduates who understand site routines, reporting, safety, and coordination between office and field teams.",
            "applications": ["Site diaries", "Field reports", "Coordination with contractors and consultants"],
            "careers": track_careers,
        }

    if kind == "elective":
        return {
            "description": "A choice-based requirement that lets students focus on an approved civil engineering or supporting technical topic.",
            "skills": ["Specialized study", "Academic planning", "Applied project thinking"],
            "jobMarket": "Electives help students build a targeted profile for internships, graduation projects, and early-career roles.",
            "applications": ["Specialized design work", "Software-supported analysis", "Applied project studies"],
            "careers": track_careers,
        }

    patterns = [
        (("english", "writing", "communication", "presentation"), {
            "description": "A communication course focused on clear academic and professional English for engineering settings.",
            "skills": ["Technical writing", "Presentation structure", "Professional vocabulary"],
            "jobMarket": "Clear communication supports reports, proposals, site correspondence, and client-facing engineering work.",
            "applications": ["Engineering reports", "Presentations", "Professional emails"],
            "careers": ["Site engineer", "Design engineer", "Project coordinator"],
        }),
        (("computer", "programming", "python", "machine learning"), {
            "description": "An applied computing course that builds digital problem-solving skills for civil engineering analysis and documentation.",
            "skills": ["Computational thinking", "Data handling", "Engineering software workflows"],
            "jobMarket": "Digital skills are increasingly expected in design offices, construction teams, and infrastructure analytics roles.",
            "applications": ["Automated calculations", "Modeling workflows", "Data-supported engineering decisions"],
            "careers": ["BIM engineer", "Digital construction engineer", "Structural analysis engineer"],
        }),
        (("calculus", "differential", "numerical", "algebra", "statistics", "optimization", "probability"), {
            "description": "A mathematics course supporting engineering modeling, analysis, uncertainty, and numerical calculation.",
            "skills": ["Quantitative modeling", "Problem formulation", "Analytical reasoning"],
            "jobMarket": "Strong mathematics supports structural analysis, hydraulics, geotechnical design, optimization, and data-informed decisions.",
            "applications": ["Load-response calculations", "Optimization studies", "Engineering data analysis"],
            "careers": ["Structural engineer", "Infrastructure analyst", "Planning engineer"],
        }),
        (("physics", "chemistry", "mechanics"), {
            "description": "A foundational science course for understanding physical behavior, forces, materials, motion, and engineering systems.",
            "skills": ["Scientific reasoning", "Free-body diagrams", "Material and force concepts"],
            "jobMarket": "These foundations are required for safe structural, materials, geotechnical, and construction decisions.",
            "applications": ["Load paths", "Material behavior", "Laboratory interpretation"],
            "careers": ["Civil engineer", "Structural engineer", "Materials testing engineer"],
        }),
        (("drawing", "drafting", "surveying", "gis"), {
            "description": "A spatial documentation course focused on drawings, survey data, mapping, and civil engineering representation.",
            "skills": ["Technical drawing", "Spatial interpretation", "Survey and mapping workflows"],
            "jobMarket": "Civil engineers use drawings, survey records, and GIS outputs to coordinate design, site execution, and asset records.",
            "applications": ["Site layouts", "CAD drawings", "GIS maps and survey documentation"],
            "careers": ["Surveying engineer", "Site engineer", "BIM/GIS specialist"],
        }),
        (("structural analysis", "structure analysis", "steel", "concrete", "foundation", "soil", "strengthening"), {
            "description": "A structural and geotechnical design course concerned with how civil infrastructure carries loads safely and efficiently.",
            "skills": ["Load analysis", "Design checks", "Code-aware engineering judgment"],
            "jobMarket": "Consultants, contractors, and public agencies need engineers who can evaluate buildings, foundations, and structural systems.",
            "applications": ["Building frames", "Foundation systems", "Structural rehabilitation"],
            "careers": ["Structural design engineer", "Geotechnical engineer", "Technical office engineer"],
        }),
        (("materials", "smart buildings", "smart materials", "quality control"), {
            "description": "A materials-focused course dealing with testing, quality, durability, and performance in built assets.",
            "skills": ["Material selection", "Testing interpretation", "Quality-control reasoning"],
            "jobMarket": "Materials knowledge supports safe construction, durability, sustainability, and performance monitoring.",
            "applications": ["Concrete testing", "Quality-control plans", "Smart material evaluation"],
            "careers": ["Materials engineer", "Quality control engineer", "Structural health monitoring engineer"],
        }),
        (("project management", "planning", "scheduling", "quantity", "estimation", "contracts", "commercial", "risk", "asset", "strategic", "real estate", "human resources", "operation research", "economics"), {
            "description": "A management-oriented civil engineering course focused on delivery, planning, cost, contracts, risk, and resources.",
            "skills": ["Planning and control", "Cost and quantity awareness", "Commercial decision-making"],
            "jobMarket": "Construction and infrastructure employers need engineers who can control time, budget, contracts, and stakeholder expectations.",
            "applications": ["Schedules", "Cost estimates", "Contract administration and risk registers"],
            "careers": ["Project controls engineer", "Quantity surveyor", "Construction project manager"],
        }),
        (("fluid", "hydraulic", "transportation", "highway"), {
            "description": "An infrastructure systems course covering water, transport, and service networks relevant to civil engineering projects.",
            "skills": ["System analysis", "Design calculation", "Infrastructure planning awareness"],
            "jobMarket": "Infrastructure agencies and consultants need engineers who understand transport, hydraulic, and public works systems.",
            "applications": ["Hydraulic calculations", "Highway components", "Infrastructure feasibility studies"],
            "careers": ["Transportation engineer", "Hydraulic engineer", "Infrastructure engineer"],
        }),
        (("smart cities", "intelligent systems", "bim", "building information modeling"), {
            "description": "A digital and smart-infrastructure course connecting civil engineering with data, models, and connected built environments.",
            "skills": ["Digital modeling", "Systems thinking", "Technology evaluation"],
            "jobMarket": "Modern civil engineering teams increasingly use BIM, smart-city systems, sensors, and analytics to plan and manage assets.",
            "applications": ["BIM coordination", "Smart infrastructure concepts", "Sensor-informed asset workflows"],
            "careers": ["BIM engineer", "Smart infrastructure engineer", "Digital construction specialist"],
        }),
        (("graduation project",), {
            "description": "A capstone design and research experience where students integrate civil engineering knowledge into a supervised project.",
            "skills": ["Problem definition", "Design integration", "Technical reporting"],
            "jobMarket": "Capstone work helps students demonstrate readiness for design offices, contractors, and infrastructure organizations.",
            "applications": ["Integrated design project", "Technical presentation", "Applied research or prototype work"],
            "careers": track_careers,
        }),
    ]
    for keys, profile in patterns:
        if any(key in lower for key in keys):
            return profile

    return {
        "description": f"A civil engineering course that develops applied knowledge related to {name.lower()}.",
        "skills": ["Engineering problem solving", "Technical documentation", "Applied civil engineering judgment"],
        "jobMarket": "The course supports the technical and professional foundation expected in entry-level civil engineering roles.",
        "applications": ["Design tasks", "Site coordination", "Engineering reports"],
        "careers": track_careers,
    }


def extract_course(ws, row: int, block: dict[str, Any], semester_number: int, term: str, track: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None]:
    code = clean(ws.cell(row, block["codeCol"]).value)
    name = clean(ws.cell(row, block["nameCol"]).value)
    if not code or not name or code.lower() == "code":
        return None, None

    flags = read_flags(ws, row, block)
    credits, note = assign_credits(code, name, ws.cell(row, block["creditCol"]).value)
    kind = course_type(code, name, flags)
    profile = course_profile(name, kind, track["name"])
    prereq = clean(ws.cell(row, block["prereqCol"]).value)
    if prereq == "--":
        prereq = ""

    return {
        "courseCode": code,
        "courseName": name,
        "creditHours": credits,
        "semester": semester_number,
        "semesterTitle": f"Semester {semester_number}",
        "term": term,
        "track": track["name"],
        "type": kind,
        "preRequisites": prereq,
        "description": profile["description"],
        "skills": profile["skills"],
        "jobMarket": profile["jobMarket"],
        "applications": profile["applications"],
        "careers": profile["careers"],
        "source": {"workbook": WORKBOOK_PATH.name, "sheet": track["sourceSheet"], "row": row},
    }, note


def extract_electives(ws, track: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    groups = {"majorElectives": range(61, 70), "mathElectives": range(74, 77)}
    output: dict[str, list[dict[str, Any]]] = {}
    for group, rows in groups.items():
        courses = []
        for row in rows:
            course, _ = extract_course(ws, row, RIGHT, 0, "Elective options", track)
            if course:
                course["semester"] = "Elective options"
                course["semesterTitle"] = "Elective options"
                course["type"] = "elective"
                courses.append(course)
        output[group] = courses
    return output


def extract_track(wb, track: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    ws = wb[track["sourceSheet"]]
    semesters = []
    notes = []
    semester_number = 1
    for semester_block in SEMESTER_BLOCKS:
        for side in (LEFT, RIGHT):
            term = clean(ws.cell(semester_block["titleRow"], side["termCol"]).value)
            courses = []
            for row in range(semester_block["startRow"], semester_block["endRow"] + 1):
                course, note = extract_course(ws, row, side, semester_number, term, track)
                if course:
                    courses.append(course)
                if note:
                    notes.append(note)
            semesters.append({
                "semester": semester_number,
                "title": f"Semester {semester_number}",
                "term": term,
                "year": semester_block["year"],
                "totalCreditHours": sum(course["creditHours"] for course in courses),
                "courses": courses,
            })
            semester_number += 1

    courses = [course for semester in semesters for course in semester["courses"]]
    total = sum(course["creditHours"] for course in courses)
    zero_credit = sum(1 for course in courses if course["creditHours"] == 0)
    type_counts = Counter(course["type"] for course in courses)
    return {
        **track,
        "semesters": semesters,
        "electiveOptions": extract_electives(ws, track),
        "totalCreditHours": total,
        "numberOfSemesters": len(semesters),
        "zeroCreditRequirements": zero_credit,
        "typeCounts": dict(type_counts),
        "validation": {"expectedTotalCreditHours": 144, "equalsExpectedTotal": total == 144},
    }, notes


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    tracks = []
    notes = []
    for track in TRACKS:
        extracted, track_notes = extract_track(wb, track)
        tracks.append(extracted)
        notes.extend(track_notes)

    payload = {
        "program": "Civil Engineering Program / CCE Department",
        "institution": "Nile University",
        "sourceWorkbook": WORKBOOK_PATH.name,
        "sheetsRead": wb.sheetnames,
        "publishedSheets": [track["sourceSheet"] for track in TRACKS],
        "tracks": tracks,
        "notes": [
            "Official course descriptions were not present in the workbook; descriptions are concise generic summaries based on course titles.",
            "The workbook includes 7-semester variants; the website publishes the 8-semester plans because they directly validate to 144 credit hours.",
            "The user rule mentions Writing Skills and Presentation as 3 credit hours; the workbook lists EGL213 Writing Skills and EGL214 Communication and Presentation Skills separately, both with 3 credit hours.",
            *notes,
        ],
    }

    CONTENT_DIR.mkdir(parents=True, exist_ok=True)
    (CONTENT_DIR / "study-plan.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    summary = {
        track["id"]: {
            "name": track["name"],
            "sourceSheet": track["sourceSheet"],
            "semesters": track["numberOfSemesters"],
            "totalCreditHours": track["totalCreditHours"],
            "zeroCreditRequirements": track["zeroCreditRequirements"],
            "equals144": track["validation"]["equalsExpectedTotal"],
        }
        for track in tracks
    }
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
